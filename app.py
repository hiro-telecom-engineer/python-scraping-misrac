import datetime
import shutil

from openpyxl import load_workbook
import openpyxl

import urllib
import urllib.request
from bs4 import BeautifulSoup


# グローバル変数
now = datetime.datetime.now()
g_file_name = '出力結果' + now.strftime('%Y%m%d_%H%M%S') + '.xlsx'	# 出力ファイル名
g_file_path = './出力/' + g_file_name									# 出力ファイルパス
g_ws = ""
g_line_num = 4

# エクセルファイル作成
shutil.copy("./misra取り込みベース.xlsx", g_file_path)
# 書き込みエクセル
g_wb = openpyxl.load_workbook(g_file_path)
g_ws = g_wb["一覧"]

# メイン関数
def main():
	global g_ws
	get_data()
	g_wb.save(g_file_path)

# 一覧更新
def write_data( misra_item , misra_data , href_url ):
	global g_ws
	global g_line_num
	g_ws.cell(row=g_line_num,column=2).value = misra_item
	counter = 3
	for data in misra_data:
		g_ws.cell(row=g_line_num,column=counter).value = data
		counter += 1
	g_ws.cell(row=g_line_num,column=counter).value = href_url
	g_line_num += 1

# データの取得
def get_data():
	#対象のURLとcsvを保存するパスを指定
	url ="http://www.c-lang.org/detail/misra_c.html"

	#urlからhtmlを取得
	html = urllib.request.urlopen(url)

	#構文解析
	soup = BeautifulSoup(html.read(),"lxml")

	#構文解析したデータからtable要素でclass属性=table_typeである部分を抽出
	tables = soup.findAll("table", {"class":"table_type"})
	csv_header = []

	#thead要素の中のth要素をfor文で取得していく
	for head in tables[0].find_all(['thead'])[0].find_all(['th']):
		#csv_headerに抜き出したデータを格納
		csv_header.append(head.get_text())

	for table in tables:
		rows = table.find_all("tr")
		caption = table.find("caption")
		for row in rows:
			misra_data =[]
			for cell in row.find_all(['td']):
				misra_data.append(cell.get_text())

			#URLがあれば取得
			href_url = row.find('a', href=True)
			if href_url:
				href_url = href_url.get("href")

			#余計な空白を除去
			if any(misra_data):
				if href_url:
					write_data(caption.get_text()[1:-1], misra_data , url + str(href_url))
				else:
					write_data(caption.get_text()[1:-1], misra_data , "-")
main()

